import io

from django.utils.translation import gettext
from django.http import HttpResponse

from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

__doc__ = """Repeated pieces of code for generating reports in Excel were placed in this file."""


def styles_and_borders():
    """First block for making excel report

    """

    # TODO: in the end of rendering excel ???
    f_response = io.BytesIO()
    # start EXCEL(styles)

    m_border = Border(
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
    )

    m_cel_item_itog = PatternFill(
        fill_type="solid",
        fgColor='00808080'
    )
    m_cel_report_itog = PatternFill(
        fill_type="solid",
        fgColor='00C0C0C0'
    )

    return f_response, m_border, m_cel_item_itog, m_cel_report_itog


def title_and_date_range(ws, row_item, title, merged_cells_title, merged_cells_date_range, add_param):
    """Second block for making excel report

    """

    # HEADER
    m_title = gettext(title)
    ws[f'A{row_item}'] = m_title

    ws.merge_cells(merged_cells_title)
    ws[f'A{row_item}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f'A{row_item}'].font = Font(bold=True)

    row_item += 1
    ws[f'A{row_item}'] = add_param.get('date_range', "")
    ws.merge_cells(merged_cells_date_range)
    ws['A{}'.format(row_item)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A{}'.format(row_item)].font = Font(bold=True)

    return ws, row_item


def making_titled_cells(ws, row_item, m_border, m_cel_item_itog, list_of_column_titles):
    """Third block for making excel report

    """

    titled_cells = []
    current_letter_index = 0
    english_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                   'U', 'V', 'W', 'X', 'Y', 'Z']

    for title in list_of_column_titles:
        titled_cells.append(gettext(title))

    row_item += 2
    for title_of_cell in titled_cells:

        cel = ws.cell(row=row_item, column=current_letter_index+1, value=title_of_cell)
        cel.alignment = Alignment(horizontal='center', vertical='center')
        cel.border = m_border
        ws.column_dimensions[english_letters[current_letter_index]].width = 20
        cel.fill = m_cel_item_itog
        current_letter_index += 1

    return ws, row_item


def row_of_cells_with_data(ws, row_item, m_border, values_for_cells):
    for index, cell in enumerate(values_for_cells):
        row_item += 1
        cel = ws.cell(row=row_item, column=index + 1, value=cell)
        cel.alignment = Alignment(horizontal='center', vertical='center')
        cel.border = m_border
    return ws, row_item


def block_separator_row(ws, row_item, m_cel_item_itog, count_of_cells):
    row_item += 1
    if count_of_cells > 1:
        cel = ws.cell(row=row_item, column=1, value="")
        cel.border = Border(
            left=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel.fill = m_cel_item_itog

        for cell in range(count_of_cells - 2):
            cel = ws.cell(row=row_item, column=cell + 2, value="")
            cel.border = Border(
                top=Side(border_style="thin", color='FF000000'),
                bottom=Side(border_style="thin", color='FF000000'),
            )
            cel.fill = m_cel_item_itog

        cel = ws.cell(row=row_item, column=count_of_cells, value="")
        cel.border = Border(
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel.fill = m_cel_item_itog

    elif count_of_cells == 1:
        cel = ws.cell(row=row_item, column=1, value="")
        cel.border = Border(
            left=Side(border_style="thin", color='FF000000'),
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel.fill = m_cel_item_itog
    return ws, row_item


def block_amount_row(ws, row_item, m_cel_item_itog, count_of_cells, amount_value):
    row_item += 1
    if count_of_cells > 1:
        cel_text = ws.cell(row=row_item, column=1, value="Amount of block:")
        cel_text.alignment = Alignment(horizontal='center', vertical='center')
        cel_text.border = Border(
            left=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel_text.fill = m_cel_item_itog

        for cell in range(count_of_cells-2):
            cel = ws.cell(row=row_item, column=cell+2, value="")
            cel.border = Border(
                top=Side(border_style="thin", color='FF000000'),
                bottom=Side(border_style="thin", color='FF000000'),
            )
            cel.fill = m_cel_item_itog

        cel_count = ws.cell(row=row_item, column=count_of_cells, value=amount_value)
        cel_count.alignment = Alignment(horizontal='center', vertical='center')
        cel_count.border = Border(
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel_count.fill = m_cel_item_itog

    elif count_of_cells == 1:
        cel = ws.cell(row=row_item, column=1, value=f"Amount of block: {amount_value}")
        cel.border = Border(
            left=Side(border_style="thin", color='FF000000'),
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel.fill = m_cel_item_itog
    return ws, row_item


def all_amount_row(ws, row_item, m_cel_report_itog, count_of_cells, all_amount_value):
    row_item += 1

    if count_of_cells > 1:
        cel_text = ws.cell(row=row_item, column=1, value="All amount:")
        cel_text.alignment = Alignment(horizontal='center', vertical='center')
        cel_text.border = Border(
            left=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel_text.fill = m_cel_report_itog

        for cell in range(count_of_cells - 2):
            cel = ws.cell(row=row_item, column=cell + 2, value="")
            cel.border = Border(
                top=Side(border_style="thin", color='FF000000'),
                bottom=Side(border_style="thin", color='FF000000'),
            )
            cel.fill = m_cel_report_itog

        cel_count = ws.cell(row=row_item, column=4, value=all_amount_value)
        cel_count.alignment = Alignment(horizontal='center', vertical='center')
        cel_count.border = Border(
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel_count.fill = m_cel_report_itog

    elif count_of_cells == 1:
        cel = ws.cell(row=row_item, column=1, value=f"All amount: {all_amount_value}")
        cel.border = Border(
            left=Side(border_style="thin", color='FF000000'),
            right=Side(border_style="thin", color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
        )
        cel.fill = m_cel_report_itog
    return ws, row_item