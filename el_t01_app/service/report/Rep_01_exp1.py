#! /usr/bin/env python
# -*- coding: utf-8 -*-
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
from django.utils.translation import gettext
from django.http import HttpResponse


def report_export(l_data_to_report, add_param):
    f_response = io.BytesIO()
    # start EXCEL
    # Stily
    m_align_horizontal = 'center'
    m_border = Border(
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
    )
    m_cel_fill = PatternFill(
        fill_type="solid",
        fgColor='FFFFFFFF'
    )
    m_cel_item_itog = PatternFill(
        fill_type="solid",
        fgColor='00C0C0C0'
    )
    m_cel_report_itog = PatternFill(
        fill_type="solid",
        fgColor='00339966'
    )
    wb = Workbook()
    ws = wb.active
    # TITLE
    row_item = 1
    m_title = gettext("Payment")
    ws[f'A{row_item}'] = m_title
    ws.merge_cells('A1:E1')
    ws[f'A{row_item}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f'A{row_item}'].font = Font(bold=True)
    row_item = 2
    ws[f'A{row_item}'] = add_param.get('date_range', "")
    ws.merge_cells('A2:E2')
    ws['A{}'.format(row_item)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A{}'.format(row_item)].font = Font(bold=True)
    # HEADER
    m_cell_01 = gettext("Payment")
    m_cell_02 = gettext("Name Game")
    m_cell_03 = gettext("Games/Transaction")
    m_cell_04 = gettext("Tickets")
    m_cell_05 = gettext("Amounts")
    row_item += 2
    # 1
    cel = ws.cell(row=row_item, column=1, value=m_cell_01)
    cel.alignment = Alignment(horizontal='center', vertical='center')
    cel.border = m_border
    ws.column_dimensions['A'].width = 20
    cel.fill = m_cel_item_itog
    # 2
    cel = ws.cell(row=row_item, column=2, value=m_cell_02)
    cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cel.border = m_border
    ws.column_dimensions['B'].width = 20
    cel.fill = m_cel_item_itog
    # 3
    cel = ws.cell(row=row_item, column=3, value=m_cell_03)
    cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cel.border = m_border
    ws.column_dimensions['C'].width = 20
    cel.fill = m_cel_item_itog
    # 4
    cel = ws.cell(row=row_item, column=4, value=m_cell_04)
    cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cel.border = m_border
    ws.column_dimensions['D'].width = 20
    cel.fill = m_cel_item_itog
    # 5
    cel = ws.cell(row=row_item, column=5, value=m_cell_05)
    cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cel.border = m_border
    ws.column_dimensions['E'].width = 20
    cel.fill = m_cel_item_itog

    for item in l_data_to_report:
        row_item += 1
        m_cell_01 = item.get("games_id__t_payment__verbal", "")
        m_cell_02 = item.get("game_type__caption", "")
        m_cell_03 = item.get("games_count", "")
        m_cell_04 = item.get("tickets_count", "")
        m_cell_05 = item.get("amount_count", "")
        # 1
        cel = ws.cell(row=row_item, column=1, value=m_cell_01)
        cel.alignment = Alignment(horizontal='center', vertical='center')
        cel.border = m_border
        if m_cell_02 == 'Total' and m_cell_01 != "Total":
            cel.fill = m_cel_report_itog
        elif m_cell_01 == "Total":
            cel.fill = m_cel_item_itog
        else:
            cel.fill = m_cel_fill
        # 2
        cel = ws.cell(row=row_item, column=2, value=m_cell_02)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border = m_border
        if m_cell_02 == 'Total' and m_cell_01 != "Total":
            cel.fill = m_cel_report_itog
        elif m_cell_01 == "Total":
            cel.fill = m_cel_item_itog
        else:
            cel.fill = m_cel_fill
        # 3
        cel = ws.cell(row=row_item, column=3, value=m_cell_03)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border = m_border
        if m_cell_02 == 'Total' and m_cell_01 != "Total":
            cel.fill = m_cel_report_itog
        elif m_cell_01 == "Total":
            cel.fill = m_cel_item_itog
        else:
            cel.fill = m_cel_fill
        # 4 Date
        cel = ws.cell(row=row_item, column=4, value=m_cell_04)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border = m_border
        if m_cell_02 == 'Total' and m_cell_01 != "Total":
            cel.fill = m_cel_report_itog
        elif m_cell_01 == "Total":
            cel.fill = m_cel_item_itog
        else:
            cel.fill = m_cel_fill
        # 5
        cel = ws.cell(row=row_item, column=5, value=m_cell_05)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border = m_border
        if m_cell_02 == 'Total' and m_cell_01 != "Total":
            cel.fill = m_cel_report_itog
        elif m_cell_01 == "Total":
            cel.fill = m_cel_item_itog
        else:
            cel.fill = m_cel_fill

    # finish export
    wb.save(f_response)
    f_response.seek(0)
    response = HttpResponse(f_response.read(), content_type="application/vnd.ms-excel")
    response['Content-Disposition'] = f'filename="report.xlsx"'
    f_response.close()
    return response
