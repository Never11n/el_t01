# ! /usr/bin/env python
# -*- coding: utf-8 -*-
import os
import json
import time
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.db.models import Q, Count

from el_t01_app.service.service import paginate
from el_t01_app.service.service import get_main_args, check_template_exists
from el_t01_app.models import Profile, Task_history, Task_status, Task_type


def cabs_task(request, cabs_type=None, task_type=None):
    print("cabs_task")
    print("cabs_type = ", cabs_type)
    args = get_main_args(request, section="main")
    if not args['logged_in']:
        return redirect('/')
    else:
        if args['me'].profile.is_boss or args['me'].profile.is_manager or args['me'].profile.is_moderator:
            args['l_head_list'] = []
            args['l_header_list'] = []
            args['l_footer_list'] = []
            args['l_script_list'] = []
            args['l_modal_list'] = []
            args['l_body_list'] = []
            args['l_layout_file'] = check_template_exists("cab_boss/layout.html")
            args['l_head_list'].append(check_template_exists("cab_boss/l_head_01.html"))
            args['l_header_list'].append(check_template_exists("cab_boss/l_header.html"))
            args['l_script_list'].append(check_template_exists("cab_boss/l_script_01.html"))
            # args['l_layout_file'] = check_template_exists("cabs_task/layout.html")
            # args['l_head_list'].append(check_template_exists("cabs_task/l_head_01.html"))
            # args['l_header_list'].append(check_template_exists("cabs_task/l_header.html"))
            # args['l_script_list'].append(check_template_exists("cabs_task/l_script_01.html"))
            if cabs_type == "cabs_list":
                m_uName = request.GET.get('uName', "")
                m_uEmail = request.GET.get('uEmail', "")
                m_uPhone = request.GET.get('uPhone', "")
                _task_list = Task_history.objects.all().order_by("-dt_add")
                args['m_addTitle'] = ""
                if m_uName != "":
                    print ("m_uName 222")
                    args['m_uName'] = m_uName
                    args['m_addTitle'] = f"name={m_uName}"
                    _task_list = _task_list.filter(name__icontains=m_uName)
                if m_uEmail != "":
                    args['m_uEmail'] = m_uEmail
                    args['m_addTitle'] += f" email={m_uName}"
                    _task_list = _task_list.filter(user__email__icontains=m_uEmail)
                if m_uPhone != "":
                    args['m_uPhone'] = m_uPhone
                    args['m_addTitle'] += f" phone={m_uPhone}"
                    _task_list = _task_list.filter(mobile__icontains=m_uPhone)
                m_list_per_page = args['GL_CabbNumberInPage']
                p = 1
                if 'page' in request.GET:
                    p = int(request.GET['page'])
                p, args['task_list'] = paginate(_task_list, p, m_list_per_page)
                args['pagin_list_all'] = _task_list.count()
                args['pagin_list_01'] = (m_list_per_page * p) - m_list_per_page + 1
                args['pagin_list_02'] = m_list_per_page * p
                if args['pagin_list_02'] > args['pagin_list_all']:
                    args['pagin_list_02'] = args['pagin_list_all']

                args['pagin_url'] = "/cabs_list"
                args['pagin_list'] = args['task_list']
                args['menu_item'] = 'nav-link_cabs_list'
                args['l_pagination'] = f"{args['GL_MainSkin']}/common/_pagination.html"

                # args['l_find'] = f"{args['GL_MainSkin']}/cabs_task/_cab_find_01.html"
                args['l_body_list'].append(check_template_exists("cab_boss_task/cabs_list.html"))
                # args['l_modal_list'].append(f"{args['GL_MainSkin']}/cabs_task/cabb_modal_editbalance.html")
                # args['l_modal_list'].append(f"{args['GL_MainSkin']}/cabs_task/_m_cabb_userprofileedit.html")
            elif cabs_type == "cabs_newtask":
                _list_user = Profile.objects.filter(reklama=True).order_by("name")
                _list_user = Profile.objects.all().order_by("name")
                m_list_per_page = args['GL_CabbNumberInPage']
                # args['list_user'] = _list_user
                # args['list_user_count'] = len(_list_user)
                # p = 1
                # if 'page' in request.GET:
                #     p = int(request.GET['page'])
                # p, args['list_user'] = paginate(_list_user, p, m_list_per_page)
                # args['pagin_list_all'] = _list_user.count()
                # args['pagin_list_01'] = (m_list_per_page * p) - m_list_per_page + 1
                # args['pagin_list_02'] = m_list_per_page * p
                # if args['pagin_list_02'] > args['pagin_list_all']:
                #     args['pagin_list_02'] = args['pagin_list_all']
                # args['pagin_url'] = "/cabb_games"
                # args['pagin_list'] = args['list_user']
                if task_type == 'sms':
                    args['menu_item'] = 'nav-link_cabs_newtask_sms'
                elif task_type == 'whatsapp':
                    args['menu_item'] = 'nav-link_cabs_newtask_whatsapp'
                elif task_type == 'email':
                    _list_user = _list_user.filter(mail_confirmed=True)
                    args['menu_item'] = 'nav-link_cabs_newtask_email'
                args['selected_task_type'] = task_type
                args['l_body_list'].append(check_template_exists("cab_boss_task/cabs_newtask.html"))
            else:  # cab_type == "boss_cabinet":
                args['menu_item'] = 'nav-link_cabinet'
                args['l_body_list'].append(check_template_exists("cab_boss/cab_game_list.html"))
            args['main_tab'] = f"{args['GL_MainSkin']}/cabinet/cabinet.html"
            return render(request, args['main_tab'], args)
        else:
            return redirect('/')


def cabs_task_users_list(request):
    args = get_main_args(request, section="main")
    if not args['logged_in'] or not args['me'].profile.is_moderator:
        return redirect('/')
    task_type = request.GET.get('f_task_type') if request.method == "GET" else request.POST.get('f_task_type')
    users_filter = request.GET.get('f_users_filter') if request.method == "GET" else request.POST.get('f_users_filter')
    days_not_played = request.GET.get('f_days_not_played')
    list_users = None
    users_emails = []
    if request.method == "POST" and 'f_file' in request.FILES:
        file = request.FILES['f_file']
        wb = load_workbook(file)
        for worksheets in wb.sheetnames:
            worksheet = wb[worksheets]
            column_a = worksheet['A']
            for cell in column_a:
                if cell.value:
                    users_emails.append(cell.value)
    if users_filter == 'all':
        list_users = Profile.objects.filter(reklama=True).order_by("name")
    elif users_filter == 'never_played':
        list_users = Profile.objects.filter(reklama=True).annotate(
            num_tickets=Count('user__game_user__games_id_history')).filter(num_tickets=0)
    elif users_filter == 'days_not_played':
        today = datetime.datetime.now()
        delta_days = int(days_not_played)
        start_day = today + relativedelta(days=-delta_days, hour=0, minute=0, second=0, microsecond=0)
        list_users = Profile.objects.filter(reklama=True).annotate(
            num_tickets=Count('user__game_user__games_id_history',
                              filter=Q(user__game_user__games_id_history__req_dt__date__range=[start_day.date(), today.date()]))
        ).filter(num_tickets=0)
    elif users_filter == 'from_file':
        list_users = Profile.objects.filter(reklama=True, user__email__in=users_emails)
    if task_type == 'email':
        list_users = list_users.filter(mail_confirmed=True).annotate()
    args['list_user'] = list_users
    args['list_user_count'] = len(list_users)
    template_name = check_template_exists("cab_boss_task/users_table.html")
    m_template = get_template(template_name)
    return HttpResponse(m_template.render(args))


def cabs_newtask_save(request):
    print ("cabs_newtask_save")
    args = get_main_args(request, section="main")
    if not args['logged_in']:
        return redirect('/')
    if not args['me'].profile.is_moderator:
        return redirect('/')

    if request.POST:
        print ("POST = ", request.POST)
        m_task_name = request.POST.get('f_task_name', "")
        m_task_type = request.POST.get('f_task_type', "")
        m_task_text = request.POST.get('f_task_text', "")
        m_task_subject = request.POST.get('f_task_subject', "")
        m_task_users = request.POST.get('f_task_users', "")
        m_task_users = m_task_users.strip().split(" ")
        m_task_dated = request.POST.get('f_task_dated', "")
        m_task_dateh = request.POST.get('f_task_dateh', "")
        m_task_datem = request.POST.get('f_task_datem', "")
        m_task_status_id = Task_status.objects.get(verbal="add").id
        m_task_type_id = Task_type.objects.get(verbal=m_task_type).id
        try:
            t_task = Task_history()
            t_task.t_status_id = m_task_status_id
            t_task.t_type_id = m_task_type_id
            t_task.task_name = m_task_name
            t_task.task_user = m_task_users
            t_task.task_num = len(m_task_users)
            t_task.task_text = m_task_text
            t_task.subject = m_task_subject
            t_task.task_add_conf1 = True
            t_task.dt_add = datetime.datetime.now()
            m_dt_plan = f'{m_task_dated} {m_task_dateh}:{m_task_datem}'
            m_dt_plan = datetime.datetime.strptime(m_dt_plan, "%Y-%m-%d %H:%M")
            t_task.dt_plan = m_dt_plan
            if 'photo1' in request.FILES:
                t_task.image = request.FILES['photo1']
            t_task.save()
            return_data = {"AnswerCod": "01"}
        except Exception as ex:
            print(f"ex = {ex}")
            print("save 2")
            return_data = {"AnswerCod": "00"}
    else:
        print("save 3")
        return_data = {"AnswerCod": "00"}
    return JsonResponse(return_data)


def cabs_nosms(request, verbal=None):
    print("cabs_nosms")
    print("verbal = ", verbal)
    args = get_main_args(request, section="main")
    # args['menu_item'] = 'nav-link_cabinet'


    if not args['logged_in']:
        return redirect('/')
