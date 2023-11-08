#! /usr/bin/env python
# -*- coding: utf-8 -*-
import os
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
from django.utils.translation import gettext
from django.http import HttpResponse

letters = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M',
           14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}


def get_nested_attr(obj, attrs, counter, depth):
    if counter == depth - 1:
        return getattr(obj, attrs[counter])
    else:
        obj = getattr(obj, attrs[counter])
        return get_nested_attr(obj, attrs, counter + 1, depth)


def boss_page_export(columns, values, queryset, title):
    f_response = io.BytesIO()
    # start EXCEL
    # Stily
    m_align_horizontal = 'center'
    m_border = Border(
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
    )
    m_cel_fill = PatternFill(
        fill_type="solid",
        fgColor='FFFFFFFF'
    )
    m_cel_item_itog = PatternFill(
        fill_type="solid",
        fgColor='00C0C0C0'
    )
    m_cel_report_itog = PatternFill(
        fill_type="solid",
        fgColor='00339966'
    )
    wb = Workbook()
    ws = wb.active
    # TITLE
    row_item = 1
    m_title = gettext(title)
    ws[f'A{row_item}'] = m_title
    columns_len = len(columns)
    ws.merge_cells(f'A1:{letters[columns_len]}1')
    ws[f'A{row_item}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f'A{row_item}'].font = Font(bold=True)
    row_item = 4
    # HEADER
    for idx, column in enumerate(columns, 1):
        cel = ws.cell(row=row_item, column=idx, value=gettext(column))
        cel.alignment = Alignment(horizontal='center', vertical='center')
        cel.border = m_border
        column_letter = letters[idx]
        ws.column_dimensions[column_letter].width = 20
    # TABLE
    for item in queryset:
        row_item += 1
        for key, val in values.items():
            cell = ''
            if type(val) == str:
                if hasattr(item, val):
                    cell = getattr(item, val)
            elif type(val) == list and len(val) > 0:
                try:
                    cell = get_nested_attr(item, val, counter=0, depth=len(val))
                except:
                    cell = ''
            elif type(val) == dict:
                for key2, val2 in val.items():
                    cell += f'{key2}: '
                    cell2 = ''
                    if type(val2) == list and len(val2) > 0:
                        try:
                            cell2 = get_nested_attr(item, val2, counter=0, depth=len(val2))
                        except:
                            cell2 = ''
                    else:
                        if hasattr(item, val2):
                            cell2 = getattr(item, val2)
                    if isinstance(cell2, datetime.datetime):
                        cell2 = datetime.datetime.strftime(cell2, "%d.%m.%Y %H:%M")
                    elif isinstance(cell2, datetime.date):
                        cell2 = f'{cell2.day}.{cell2.month}.{cell2.year}'
                    cell += f'{cell2} \n'
            if isinstance(cell, datetime.datetime):
                cell = datetime.datetime.strftime(cell, "%d.%m.%Y %H:%M")
            elif isinstance(cell, datetime.date):
                cell = f'{cell.day}.{cell.month}.{cell.year}'
            cel = ws.cell(row=row_item, column=key, value=cell)
            cel.alignment = Alignment(horizontal='center', vertical='center')
            cel.border = m_border
            cel.fill = m_cel_fill
    # finish export
    wb.save(f_response)
    f_response.seek(0)
    response = HttpResponse(f_response.read(), content_type="application/vnd.ms-excel")
    response['Content-Disposition'] = f'filename="file.xlsx"'
    f_response.close()
    return response
